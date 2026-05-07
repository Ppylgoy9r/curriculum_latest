import os
import json
from django.http import JsonResponse
from django.views import View
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import default_storage
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import Batch, Curriculum, Analysis
from .serializers import (
    BatchSerializer, CurriculumSerializer, CurriculumUploadSerializer,
    AnalysisSerializer, AnalysisSummarySerializer
)
from .ai_analyzer import parse_xlsx, analyze_curriculum_with_ai


@method_decorator(csrf_exempt, name='dispatch')
class BatchListCreateView(APIView):
    """List all batches or create a new batch."""
    def get(self, request):
        batches = Batch.objects.all()
        serializer = BatchSerializer(batches, many=True)
        return Response(serializer.data)

    def post(self, request):
        serializer = BatchSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@method_decorator(csrf_exempt, name='dispatch')
class BatchDetailView(APIView):
    """Retrieve, update or delete a batch."""
    def get(self, request, pk):
        try:
            batch = Batch.objects.get(pk=pk)
        except Batch.DoesNotExist:
            return Response({'error': 'Batch not found'}, status=status.HTTP_404_NOT_FOUND)
        serializer = BatchSerializer(batch)
        return Response(serializer.data)

    def put(self, request, pk):
        try:
            batch = Batch.objects.get(pk=pk)
        except Batch.DoesNotExist:
            return Response({'error': 'Batch not found'}, status=status.HTTP_404_NOT_FOUND)
        serializer = BatchSerializer(batch, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        try:
            batch = Batch.objects.get(pk=pk)
        except Batch.DoesNotExist:
            return Response({'error': 'Batch not found'}, status=status.HTTP_404_NOT_FOUND)
        batch.delete()
        return Response({'message': 'Batch deleted'}, status=status.HTTP_204_NO_CONTENT)


@method_decorator(csrf_exempt, name='dispatch')
class CurriculumUploadView(APIView):
    """Upload and parse curriculum XLSX file."""
    def post(self, request):
        if 'file' not in request.FILES:
            return Response({'error': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)

        uploaded_file = request.FILES['file']
        file_name = uploaded_file.name
        file_size = uploaded_file.size

        # Validate file extension
        valid_extensions = ['.xlsx', '.xls', '.csv', '.pdf', '.doc', '.docx']
        ext = os.path.splitext(file_name)[1].lower()
        if ext not in valid_extensions:
            return Response(
                {'error': f'Invalid file type. Supported: {", ".join(valid_extensions)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validate file size (50MB max)
        if file_size > 52428800:
            return Response({'error': 'File too large. Maximum size is 50MB.'}, status=status.HTTP_400_BAD_REQUEST)

        # Save file
        file_path = default_storage.save(f'curricula/{file_name}', uploaded_file)
        full_path = default_storage.path(file_path)

        # Get or create batch
        batch = None
        batch_name = request.data.get('batch_name', '')
        batch_year = request.data.get('batch_year', '2025-2026')
        batch_id = request.data.get('batch_id')

        if batch_id:
            try:
                batch = Batch.objects.get(pk=batch_id)
            except Batch.DoesNotExist:
                pass

        if not batch and batch_name:
            batch, created = Batch.objects.get_or_create(
                name=batch_name,
                defaults={'year': batch_year}
            )

        # Parse XLSX file
        weeks_data = {}
        curriculum_text = ''
        if ext in ['.xlsx', '.xls']:
            try:
                weeks_data = parse_xlsx(full_path)
                # Build curriculum text from weeks_data
                parts = []
                for week, data in weeks_data.items():
                    parts.append(f"{week}: {', '.join(data.get('topics', []))}")
                curriculum_text = '\n'.join(parts)
            except Exception as e:
                return Response({'error': f'Failed to parse file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            # For non-XLSX files, store basic info
            curriculum_text = f"File: {file_name}"

        # Create curriculum record
        curriculum = Curriculum.objects.create(
            batch=batch,
            file_name=file_name,
            file_size=file_size,
            file_path=file_path,
            weeks_data=weeks_data
        )

        return Response({
            'id': str(curriculum.id),
            'file_name': file_name,
            'file_size': file_size,
            'weeks_data': weeks_data,
            'batch': str(batch.id) if batch else None,
            'curriculum_text': curriculum_text,
            'message': 'File uploaded successfully'
        }, status=status.HTTP_201_CREATED)


@method_decorator(csrf_exempt, name='dispatch')
class CurriculumListView(APIView):
    """List all curricula."""
    def get(self, request):
        curricula = Curriculum.objects.all()
        serializer = CurriculumSerializer(curricula, many=True)
        return Response(serializer.data)


@method_decorator(csrf_exempt, name='dispatch')
class AnalyzeCurriculumView(APIView):
    """Analyze a curriculum using AI."""
    def post(self, request, curriculum_id):
        try:
            curriculum = Curriculum.objects.get(pk=curriculum_id)
        except Curriculum.DoesNotExist:
            return Response({'error': 'Curriculum not found'}, status=status.HTTP_404_NOT_FOUND)

        # Build curriculum text
        curriculum_text = ''
        if curriculum.weeks_data:
            parts = []
            for week, data in curriculum.weeks_data.items():
                topics = data.get('topics', [])
                if topics:
                    parts.append(f"{week}: {', '.join(topics)}")
                else:
                    parts.append(f"{week}: {data.get('raw_text', '')}")
            curriculum_text = '\n'.join(parts)

        # Perform AI analysis
        analysis_data = analyze_curriculum_with_ai(curriculum.weeks_data, curriculum_text)

        # Update report summary with actual counts
        total_weeks = len(curriculum.weeks_data) if curriculum.weeks_data else 0
        total_topics = sum(len(data.get('topics', [])) for data in curriculum.weeks_data.values()) if curriculum.weeks_data else 0

        if 'report_summary' in analysis_data:
            analysis_data['report_summary']['total_weeks'] = total_weeks
            analysis_data['report_summary']['total_topics'] = total_topics

        # Create analysis record
        analysis = Analysis.objects.create(
            curriculum=curriculum,
            overall_score=analysis_data.get('overall_score', 0),
            confidence_score=analysis_data.get('confidence_score', 0),
            report_status=analysis_data.get('report_status', 'pending'),
            detected_topic=analysis_data.get('detected_topic', ''),
            subtopics=analysis_data.get('subtopics', []),
            whats_good=analysis_data.get('whats_good', []),
            needs_improvement=analysis_data.get('needs_improvement', []),
            ai_suggestions=analysis_data.get('ai_suggestions', []),
            quick_fixes=analysis_data.get('quick_fixes', []),
            category_scores=analysis_data.get('category_scores', {}),
            score_trend=analysis_data.get('score_trend', []),
            improvement_summary=analysis_data.get('improvement_summary', {}),
            previous_score=analysis_data.get('previous_score', 0),
            performance_change=analysis_data.get('performance_change', 0.0),
            categories_improved=analysis_data.get('categories_improved', 0),
            categories_total=analysis_data.get('categories_total', 0),
            full_report=analysis_data
        )

        serializer = AnalysisSummarySerializer(analysis)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


@method_decorator(csrf_exempt, name='dispatch')
class AnalysisDetailView(APIView):
    """Get analysis details."""
    def get(self, request, analysis_id):
        try:
            analysis = Analysis.objects.get(pk=analysis_id)
        except Analysis.DoesNotExist:
            return Response({'error': 'Analysis not found'}, status=status.HTTP_404_NOT_FOUND)
        serializer = AnalysisSummarySerializer(analysis)
        return Response(serializer.data)


@method_decorator(csrf_exempt, name='dispatch')
class AnalysisListView(APIView):
    """List all analyses."""
    def get(self, request):
        analyses = Analysis.objects.all()[:20]
        serializer = AnalysisSummarySerializer(analyses, many=True)
        return Response(serializer.data)


@method_decorator(csrf_exempt, name='dispatch')
class DownloadAnalysisView(View):
    """Download analysis as XLSX."""
    def get(self, request, analysis_id):
        try:
            analysis = Analysis.objects.get(pk=analysis_id)
        except Analysis.DoesNotExist:
            return JsonResponse({'error': 'Analysis not found'}, status=404)

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = Workbook()

            # Sheet 1: Overview
            ws1 = wb.active
            ws1.title = "Overview"

            header_font = Font(bold=True, size=14, color="FFFFFF")
            header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
            sub_header_font = Font(bold=True, size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            ws1.merge_cells('A1:D1')
            ws1['A1'] = 'Curriculum Analysis Report'
            ws1['A1'].font = Font(bold=True, size=18, color="2563EB")

            ws1['A3'] = 'File Name'
            ws1['B3'] = analysis.curriculum.file_name
            ws1['A4'] = 'Overall Score'
            ws1['B4'] = f"{analysis.overall_score}/100"
            ws1['A5'] = 'Status'
            ws1['B5'] = analysis.report_status
            ws1['A6'] = 'Confidence'
            ws1['B6'] = f"{analysis.confidence_score}%"
            ws1['A7'] = 'Detected Topic'
            ws1['B7'] = analysis.detected_topic
            ws1['A8'] = 'Analyzed At'
            ws1['B8'] = str(analysis.analyzed_at)

            for row in ws1.iter_rows(min_row=3, max_row=8, min_col=1, max_col=2):
                for cell in row:
                    cell.border = border
                row[0].font = sub_header_font

            # Sheet 2: Category Scores
            ws2 = wb.create_sheet("Category Scores")
            ws2['A1'] = 'Category'
            ws2['B1'] = 'Score'
            ws2['A1'].font = header_font
            ws2['A1'].fill = header_fill
            ws2['B1'].font = header_font
            ws2['B1'].fill = header_fill

            row_num = 2
            for category, score in analysis.category_scores.items():
                ws2[f'A{row_num}'] = category
                ws2[f'B{row_num}'] = score
                for col in ['A', 'B']:
                    ws2[f'{col}{row_num}'].border = border
                row_num += 1

            # Sheet 3: Strengths & Improvements
            ws3 = wb.create_sheet("Strengths & Improvements")
            ws3['A1'] = "What's Good"
            ws3['A1'].font = Font(bold=True, size=14, color="10B981")
            for i, item in enumerate(analysis.whats_good, start=2):
                ws3[f'A{i}'] = f"  {i-1}. {item}"

            col_start = max(3, len(analysis.whats_good) + 4)
            ws3[f'C{1}'] = "Needs Improvement"
            ws3[f'C1'].font = Font(bold=True, size=14, color="F59E0B")
            for i, item in enumerate(analysis.needs_improvement, start=2):
                ws3[f'C{i}'] = f"  {i-1}. {item}"

            # Sheet 4: AI Suggestions
            ws4 = wb.create_sheet("AI Suggestions")
            ws4['A1'] = 'Title'
            ws4['B1'] = 'Description'
            ws4['C1'] = 'Priority'
            for col in ['A', 'B', 'C']:
                ws4[f'{col}1'].font = header_font
                ws4[f'{col}1'].fill = header_fill

            for i, sug in enumerate(analysis.ai_suggestions, start=2):
                ws4[f'A{i}'] = sug.get('title', '')
                ws4[f'B{i}'] = sug.get('description', '')
                ws4[f'C{i}'] = sug.get('priority', '')
                for col in ['A', 'B', 'C']:
                    ws4[f'{col}{i}'].border = border

            # Adjust column widths
            for ws in [ws1, ws2, ws3, ws4]:
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    ws.column_dimensions[column].width = min(max_length + 4, 50)

            # Save to response
            from django.http import HttpResponse
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="analysis_report.xlsx"'
            wb.save(response)
            return response

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
