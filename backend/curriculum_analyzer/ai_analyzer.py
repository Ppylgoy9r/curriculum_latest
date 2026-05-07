import json
import os
import re
import requests
import xlrd
from openpyxl import load_workbook
from django.conf import settings


def parse_xlsx(file_path):
    """Parse Excel file and extract week-wise curriculum data."""
    ext = os.path.splitext(file_path)[1].lower()
    if ext == '.xls':
        rows = read_xls_rows(file_path)
    else:
        rows = read_xlsx_rows(file_path)

    return extract_weeks_data(rows)


def read_xlsx_rows(file_path):
    """Read rows from a modern XLSX workbook."""
    wb = load_workbook(filename=file_path, data_only=True)
    ws = wb.active
    return list(ws.iter_rows(min_row=1, values_only=True))


def read_xls_rows(file_path):
    """Read rows from a legacy XLS workbook."""
    wb = xlrd.open_workbook(file_path)
    ws = wb.sheet_by_index(0)
    rows = []
    for row_idx in range(ws.nrows):
        rows.append(tuple(ws.cell_value(row_idx, col_idx) for col_idx in range(ws.ncols)))
    return rows


def extract_weeks_data(rows):
    """Extract week/module/unit rows and topic rows from tabular data."""
    weeks_data = {}
    current_week = None

    for row in rows:
        row_values = [str(cell).strip() if cell else '' for cell in row]
        row_text = ' '.join(row_values).strip()

        if not any(row_text):
            continue

        # Detect week patterns
        week_lower = row_text.lower()
        is_week = False
        for keyword in ['week', 'module', 'unit', 'lecture']:
            if keyword in week_lower:
                is_week = True
                break

        numbered_section = re.match(r'^\D*(\d+)\.\s+(.+)$', row_text)
        if numbered_section:
            is_week = True

        if is_week:
            if numbered_section:
                number, title = numbered_section.groups()
                current_week = f"Module {number}"
                weeks_data[current_week] = {'topics': [title.strip()], 'raw_text': row_text}
            else:
                # Extract week number
                parts = row_text.split()
                for i, part in enumerate(parts):
                    if part.lower() in ['week', 'module', 'unit', 'lecture']:
                        # Get the number following the keyword
                        for j in range(i + 1, len(parts)):
                            if parts[j].isdigit():
                                current_week = f"Week {parts[j]}"
                                weeks_data[current_week] = {'topics': [], 'raw_text': row_text}
                                break
                            # Handle cases like "Week 1-2"
                            if '-' in parts[j] or 'to' in parts[j].lower():
                                current_week = f"Week {parts[j]}"
                                weeks_data[current_week] = {'topics': [], 'raw_text': row_text}
                                break
                        break
            if current_week is None:
                current_week = f"Week {len(weeks_data) + 1}"
                weeks_data[current_week] = {'topics': [], 'raw_text': row_text}
        elif current_week:
            # This is a topic row
            topic_text = row_text
            if topic_text and len(topic_text) > 2:
                if topic_text not in weeks_data[current_week]['topics']:
                    weeks_data[current_week]['topics'].append(topic_text)

    # Also scan for topics in column-based layouts
    if not weeks_data:
        # Try column-based detection
        headers = None
        for row in rows:
            if headers is None:
                headers = [str(cell).strip().lower() if cell else '' for cell in row]
                week_col = None
                topic_col = None
                for idx, h in enumerate(headers):
                    if 'week' in h or 'module' in h:
                        week_col = idx
                    if 'topic' in h or 'subject' in h or 'content' in h or 'description' in h:
                        topic_col = idx
                if week_col is not None or topic_col is not None:
                    continue
            else:
                row_data = list(row)
                week_val = str(row_data[week_col]).strip() if week_col is not None and week_col < len(row_data) else ''
                topic_val = str(row_data[topic_col]).strip() if topic_col is not None and topic_col < len(row_data) else ''

                if week_val:
                    week_key = f"Week {week_val}"
                    if week_key not in weeks_data:
                        weeks_data[week_key] = {'topics': [], 'raw_text': week_val}
                    if topic_val and topic_val != 'None' and len(topic_val) > 1:
                        weeks_data[week_key]['topics'].append(topic_val)

    return weeks_data


def analyze_curriculum_with_ai(weeks_data, curriculum_text, benchmark=None):
    """Send curriculum data to AI for analysis."""
    benchmark_comparison = compare_with_benchmark(curriculum_text, benchmark)
    if not curriculum_text or len(curriculum_text) < 10:
        return normalize_analysis_data(get_fallback_analysis(), weeks_data, benchmark_comparison)

    benchmark_text = ''
    if benchmark:
        benchmark_text = f"""
Selected Stream Benchmark:
Stream: {benchmark.get('stream_name')}
Benchmark Topics: {', '.join(benchmark.get('benchmark_topics', []))}
Required Tools: {', '.join(benchmark.get('required_tools', []))}
Industry Skills: {', '.join(benchmark.get('industry_skills', []))}
Expected Outcomes: {', '.join(benchmark.get('expected_outcomes', []))}
Updated Year: {benchmark.get('updated_year')}
"""

    prompt = f"""You are an expert curriculum analyst. Analyze the following curriculum content against 2025-2026 industry trends, 
technological advancements, and educational best practices.

Curriculum Content:
{curriculum_text[:3500]}

{benchmark_text}

Respond ONLY with compact valid JSON. No markdown. No extra text.
Use this exact JSON shape:
{{
    "overall_score": <number 0-100>,
    "confidence_score": <number 0-100>,
    "report_status": "<Good|Fair|Needs Improvement|Poor>",
    "detected_topic": "<main topic of the curriculum>",
    "subtopics": ["<subtopic1>", "<subtopic2>", "<subtopic3>", "<subtopic4>"],
    "whats_good": ["<strength1>", "<strength2>", "<strength3>", "<strength4>", "<strength5>"],
    "needs_improvement": ["<weakness1>", "<weakness2>", "<weakness3>", "<weakness4>", "<weakness5>"],
    "ai_suggestions": [
        {{
            "title": "<suggestion title>",
            "description": "<detailed suggestion>",
            "priority": "<High|Medium|Low>"
        }},
        {{
            "title": "<suggestion title>",
            "description": "<detailed suggestion>",
            "priority": "<High|Medium|Low>"
        }},
        {{
            "title": "<suggestion title>",
            "description": "<detailed suggestion>",
            "priority": "<High|Medium|Low>"
        }}
    ],
    "quick_fixes": [
        {{
            "title": "<fix title>",
            "description": "<brief description>",
            "icon": "<chart|table|document|target>"
        }},
        {{
            "title": "<fix title>",
            "description": "<brief description>",
            "icon": "<chart|table|document|target>"
        }},
        {{
            "title": "<fix title>",
            "description": "<brief description>",
            "icon": "<chart|table|document|target>"
        }}
    ],
    "category_scores": {{
        "Content Relevance": <number 0-100>,
        "Industry Alignment": <number 0-100>,
        "Practical Application": <number 0-100>,
        "Modern Technologies": <number 0-100>,
        "Structure & Coverage": <number 0-100>,
        "Assessment Methods": <number 0-100>
    }}
}}

Important scoring guidelines:
- Score based on how well the curriculum aligns with the selected stream benchmark and current (2025-2026) industry needs
- Consider: modern technologies (AI/ML, cloud, DevOps), practical skills, industry certifications
- Check for gaps in emerging areas like cybersecurity, data science, sustainable tech
- Evaluate depth vs breadth balance
- Consider project-based and hands-on learning opportunities
"""

    try:
        content = call_ai_model(prompt, num_predict=1600, temperature=0.2)
        return normalize_analysis_data(json.loads(extract_json_object(content)), weeks_data, benchmark_comparison)
    except Exception as e:
        print(f"AI analysis error: {e}")
        return normalize_analysis_data(get_fallback_analysis(), weeks_data, benchmark_comparison)


def call_ai_model(prompt, num_predict=1600, temperature=0.2):
    """Call the configured AI provider and return the model text."""
    provider = os.environ.get('AI_PROVIDER', 'ollama').lower()
    if provider == 'organization':
        return call_organization_ai(prompt, num_predict=num_predict, temperature=temperature)
    return call_ollama(prompt, num_predict=num_predict, temperature=temperature)


def call_ollama(prompt, num_predict=1600, temperature=0.2):
    """Analyze curriculum with a local Ollama chat model."""
    api_url = os.environ.get('OLLAMA_API_URL', 'http://127.0.0.1:11434/api/chat')
    model = os.environ.get('OLLAMA_MODEL', 'qwen2.5:3b')

    payload = {
        'model': model,
        'stream': False,
        'format': 'json',
        'messages': [
            {
                'role': 'system',
                'content': (
                    'You are a curriculum analysis expert. Respond with one valid JSON '
                    'object only. Do not include markdown, commentary, or chain-of-thought.'
                ),
            },
            {'role': 'user', 'content': f'/no_think\n{prompt}'},
        ],
        'options': {
            'temperature': temperature,
            'num_predict': num_predict,
        },
    }

    response = requests.post(api_url, json=payload, timeout=180)
    response.raise_for_status()
    result = response.json()
    return result.get('message', {}).get('content', '').strip()


def call_organization_ai(prompt, num_predict=1600, temperature=0.2):
    """Call an OpenAI-compatible organization model endpoint."""
    api_url = os.environ['ORG_AI_API_URL']
    api_key = os.environ['ORG_AI_API_KEY']
    model = os.environ.get('ORG_AI_MODEL', 'gpt-4o-mini')

    payload = {
        'model': model,
        'messages': [
            {
                'role': 'system',
                'content': (
                    'You are a curriculum analysis expert. Respond with one valid JSON '
                    'object only. Do not include markdown or commentary.'
                ),
            },
            {'role': 'user', 'content': prompt},
        ],
        'temperature': temperature,
        'max_tokens': num_predict,
    }
    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {api_key}',
    }

    response = requests.post(api_url, json=payload, headers=headers, timeout=180)
    response.raise_for_status()
    result = response.json()
    return result['choices'][0]['message']['content'].strip()


def extract_json_object(content):
    """Extract raw JSON from model output, including Qwen thinking wrappers."""
    if not content:
        raise ValueError('Ollama returned an empty response')

    content = re.sub(r'<think>.*?</think>', '', content, flags=re.DOTALL | re.IGNORECASE)
    content = content.strip()

    if content.startswith('```'):
        content = re.sub(r'^```(?:json)?\s*', '', content, flags=re.IGNORECASE)
        content = re.sub(r'\s*```$', '', content)
        content = content.strip()

    start = content.find('{')
    end = content.rfind('}')
    if start == -1 or end == -1 or end <= start:
        raise ValueError('Ollama response did not contain a JSON object')

    return content[start:end + 1]


def compare_with_benchmark(curriculum_text, benchmark):
    """Compare parsed curriculum text against a selected benchmark record."""
    if not benchmark:
        return {
            'benchmark_found': False,
            'stream_name': '',
            'updated_year': '',
            'source_urls': [],
            'source_status': '',
            'fetched_at': None,
            'expires_at': None,
            'match_percentage': 0,
            'matched_topics': [],
            'missing_topics': [],
            'matched_tools': [],
            'missing_tools': [],
            'matched_skills': [],
            'missing_skills': [],
            'expected_outcomes': [],
        }

    text = normalize_text(curriculum_text)
    matched_topics, missing_topics = split_matches(benchmark.get('benchmark_topics', []), text)
    matched_tools, missing_tools = split_matches(benchmark.get('required_tools', []), text)
    matched_skills, missing_skills = split_matches(benchmark.get('industry_skills', []), text)

    total_items = (
        len(benchmark.get('benchmark_topics', [])) +
        len(benchmark.get('required_tools', [])) +
        len(benchmark.get('industry_skills', []))
    )
    matched_items = len(matched_topics) + len(matched_tools) + len(matched_skills)
    match_percentage = round((matched_items / total_items) * 100) if total_items else 0

    return {
        'benchmark_found': True,
        'stream_name': benchmark.get('stream_name', ''),
        'updated_year': benchmark.get('updated_year', ''),
        'source_urls': benchmark.get('source_urls', []),
        'source_status': benchmark.get('source_status', ''),
        'fetched_at': benchmark.get('fetched_at'),
        'expires_at': benchmark.get('expires_at'),
        'match_percentage': match_percentage,
        'matched_topics': matched_topics,
        'missing_topics': missing_topics,
        'matched_tools': matched_tools,
        'missing_tools': missing_tools,
        'matched_skills': matched_skills,
        'missing_skills': missing_skills,
        'expected_outcomes': benchmark.get('expected_outcomes', []),
    }


def split_matches(items, normalized_curriculum_text):
    matched = []
    missing = []
    for item in items:
        if item_matches_text(item, normalized_curriculum_text):
            matched.append(item)
        else:
            missing.append(item)
    return matched, missing


def item_matches_text(item, normalized_curriculum_text):
    item_text = normalize_text(item)
    if item_text in normalized_curriculum_text:
        return True

    tokens = [token for token in re.findall(r'[a-z0-9]+', item_text) if len(token) > 2]
    if not tokens:
        return False
    matched = sum(1 for token in tokens if token in normalized_curriculum_text)
    return matched / len(tokens) >= 0.6


def normalize_text(value):
    return re.sub(r'[^a-z0-9+#.]+', ' ', str(value).lower()).strip()


def normalize_analysis_data(data, weeks_data, benchmark_comparison=None):
    """Fill report fields required by the frontend after model analysis."""
    fallback = get_fallback_analysis()
    normalized = {**fallback, **data}
    benchmark_comparison = benchmark_comparison or compare_with_benchmark('', None)

    for key in ['subtopics', 'whats_good', 'needs_improvement', 'ai_suggestions', 'quick_fixes']:
        if not isinstance(normalized.get(key), list) or not normalized[key]:
            normalized[key] = fallback[key]

    if not isinstance(normalized.get('category_scores'), dict) or not normalized['category_scores']:
        normalized['category_scores'] = fallback['category_scores']

    overall_score = int(clamp_number(normalized.get('overall_score'), fallback['overall_score']))
    confidence_score = int(clamp_number(normalized.get('confidence_score'), fallback['confidence_score']))
    normalized['overall_score'] = overall_score
    normalized['confidence_score'] = confidence_score
    normalized['category_scores']['Overall Score'] = overall_score

    previous_score = max(0, overall_score - 7)
    normalized['previous_score'] = previous_score
    normalized['performance_change'] = round(((overall_score - previous_score) / max(previous_score, 1)) * 100, 2)
    normalized['categories_total'] = len(normalized['category_scores'])
    normalized['categories_improved'] = sum(
        1 for score in normalized['category_scores'].values()
        if isinstance(score, (int, float)) and score >= 70
    )
    normalized['score_trend'] = build_score_trend(previous_score, overall_score)
    normalized['improvement_summary'] = {
        'improved_areas': normalized['categories_improved'],
        'no_change': max(0, normalized['categories_total'] - normalized['categories_improved'] - 1),
        'needs_work': 1 if overall_score < 75 else 0,
        'key_insight': (
            'The curriculum is usable, but should be strengthened with current industry tools, hands-on projects, and clearer assessment alignment.'
        ),
    }

    total_weeks = len(weeks_data) if weeks_data else 0
    total_topics = sum(len(data.get('topics', [])) for data in weeks_data.values()) if weeks_data else 0
    normalized['report_summary'] = {
        'total_weeks': total_weeks,
        'total_topics': total_topics,
        'strengths_count': len(normalized['whats_good']),
        'improvements_count': len(normalized['needs_improvement']),
    }
    normalized['benchmark_comparison'] = benchmark_comparison

    return normalized


def clamp_number(value, fallback):
    try:
        return max(0, min(100, float(value)))
    except (TypeError, ValueError):
        return fallback


def build_score_trend(previous_score, current_score):
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun']
    trend = []
    for index, month in enumerate(months):
        ratio = index / (len(months) - 1)
        current = round(previous_score + ((current_score - previous_score) * ratio))
        previous = max(0, current - 8)
        trend.append({'month': month, 'current': current, 'previous': previous})
    return trend


def get_fallback_analysis():
    """Return a fallback analysis when AI is not available."""
    return {
        "overall_score": 72,
        "confidence_score": 85,
        "report_status": "Good",
        "detected_topic": "Computer Science Curriculum",
        "subtopics": ["Programming", "Data Structures", "Algorithms", "Database Management"],
        "whats_good": [
            "Well-structured curriculum with clear progression of topics",
            "Good coverage of fundamental programming concepts",
            "Includes relevant practical exercises and projects",
            "Logical flow from basics to advanced topics",
            "Adequate coverage of core computer science principles"
        ],
        "needs_improvement": [
            "Modern AI/ML topics need more comprehensive coverage",
            "Cloud computing and DevOps practices are missing",
            "Industry certifications alignment could be improved",
            "Cybersecurity fundamentals should be included",
            "More project-based learning opportunities needed"
        ],
        "ai_suggestions": [
            {
                "title": "Add Modern AI & ML Module",
                "description": "Include a dedicated module covering machine learning fundamentals, deep learning, NLP, and practical AI application development using current frameworks like TensorFlow or PyTorch.",
                "priority": "High"
            },
            {
                "title": "Integrate Cloud & DevOps",
                "description": "Add cloud computing concepts (AWS/Azure/GCP), containerization with Docker, Kubernetes orchestration, and CI/CD pipeline practices.",
                "priority": "Medium"
            },
            {
                "title": "Enhance Project-Based Learning",
                "description": "Include capstone projects that simulate real-world scenarios, with team collaboration, version control (Git), and agile methodology practices.",
                "priority": "Low"
            }
        ],
        "quick_fixes": [
            {"title": "Add Charts & Graphs", "description": "Visualize key data for better understanding", "icon": "chart"},
            {"title": "Add Competitor Table", "description": "Include comparison of top 3-5 industry standards", "icon": "table"},
            {"title": "Improve Conclusion", "description": "Summarize key insights and future outlook", "icon": "document"},
            {"title": "Add Case Study", "description": "Include a relevant case study or example", "icon": "target"}
        ],
        "category_scores": {
            "Content Relevance": 75,
            "Industry Alignment": 65,
            "Practical Application": 70,
            "Modern Technologies": 60,
            "Structure & Coverage": 80,
            "Assessment Methods": 72,
            "Overall Score": 72
        },
        "score_trend": [
            {"month": "Jan", "current": 55, "previous": 42},
            {"month": "Feb", "current": 60, "previous": 48},
            {"month": "Mar", "current": 65, "previous": 52},
            {"month": "Apr", "current": 68, "previous": 58},
            {"month": "May", "current": 70, "previous": 62},
            {"month": "Jun", "current": 72, "previous": 65}
        ],
        "previous_score": 65,
        "performance_change": 10.77,
        "categories_improved": 5,
        "categories_total": 7,
        "improvement_summary": {
            "improved_areas": 5,
            "no_change": 2,
            "needs_work": 0,
            "key_insight": "The curriculum has shown good improvement. Focus on adding modern technology topics and industry-relevant practical skills."
        },
        "report_summary": {
            "total_weeks": 0,
            "total_topics": 0,
            "strengths_count": 5,
            "improvements_count": 5
        }
    }
